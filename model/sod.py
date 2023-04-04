from openpyxl import load_workbook
filename = 'Sod.xlsx'
wb = load_workbook(filename=filename)
ws = wb.active

class Sod:

    def __init__(self, codigo1, servico1, codigo2, servico2):
        self.__codigo1 = codigo1
        self.__servico1 = servico1
        self.__codigo2 = codigo2
        self.__servico2 = servico2

    def set_sod(self):
        codigo1 = self.__codigo1
        servico1 = self.__servico1
        codigo2 = self.__codigo2
        servico2 = self.__servico2
        proxima_linha = ws.max_row + 1
        ws.cell(row=proxima_linha, column=1).value = codigo1
        ws.cell(row=proxima_linha, column=2).value = servico1
        ws.cell(row=proxima_linha, column=3).value = codigo2
        ws.cell(row=proxima_linha, column=4).value = servico2
        wb.save(filename=filename)

    def get_sod(self, codigo):
        for linha in range(2, ws.max_row + 1):
            if ws.cell(row=linha, column=1).value == codigo:
                nome_servico = ws.cell(row=linha, column=2).value
                return (codigo, nome_servico)
        return None

""" # # Instancia um objeto da classe Sistema
sod = Sod('445755277', 'Serviço 1', '5465231351', 'Serviço 2')
sod1 = Sod('445755277', 'Serviço 5', '5465231351', 'Serviço 2')
sod2 = Sod('445755277', 'Serviço 4', '5465231351', 'Serviço 2')
sod3 = Sod('445755277', 'Serviço 3', '5465231351', 'Serviço 2')
sod4 = Sod('445755277', 'Serviço 1', '5465231351', 'Serviço 2')
#
# # Chama o método set_servico() para salvar o objeto na planilha do Excel
sod.set_sod() 
sod1.set_sod() 
sod2.set_sod() 
sod3.set_sod() 
sod4.set_sod() 
 """
