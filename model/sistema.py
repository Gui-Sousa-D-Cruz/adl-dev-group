from openpyxl import Workbook, load_workbook
filename = 'Sistemas.xlsx'
wb = load_workbook(filename = filename)
ws = wb.active

class Sistema:

    def __init__(self, codigo, servico=None):
        self.__codigo = codigo
        if servico is None:
            self.__servico = ""
        else:
            self.__servico = servico

    def set_servico(self):
        codigo = self.__codigo
        servico = self.__servico
        proxima_linha = ws.max_row + 1
        ws.cell(row=proxima_linha, column=1).value = codigo
        ws.cell(row=proxima_linha, column=2).value = servico
        wb.save(filename = filename)

    def get_servico(self, codigo):
        for linha in range(2, ws.max_row + 1):
            if ws.cell(row=linha, column=1).value == codigo:
                nome_servico = ws.cell(row=linha, column=2).value
                return (codigo, nome_servico)
        print('Servico não encontrado')
        return None

# # Instancia um objeto da classe Sistema
sistema = Sistema('52452652', 'Servico Teste 1')
#
# # Chama o método set_servico() para salvar o objeto na planilha do Excel
sistema.set_servico()




