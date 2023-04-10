from openpyxl import Workbook, load_workbook
filename = './model/Sistemas.xlsx'
wb = load_workbook(filename=filename)
ws = wb.active

class Sistema:



    def __init__(self, codigo, nome=None):
        self.codigo = codigo
        if nome is None:
            self.nome= ""
        else:
            self.nome = nome

    def set_sistema(self):
        if self.get_sistema():
            print('Sistema ja cadastrado!')
        else:
            codigo = self.codigo
            nome = self.nome
            proxima_linha = ws.max_row + 1
            ws.cell(row=proxima_linha, column=1).value = codigo
            ws.cell(row=proxima_linha, column=2).value = nome
            wb.save(filename = filename)

    def get_sistema(self):
        for linha in range(2, ws.max_row + 1):
            if ws.cell(row=linha, column=1).value == self.codigo:
                nome_sistema = ws.cell(row=linha, column=2).value
                return f'Sistema encontrado! \n Código => {self.codigo} \n Nome: {nome_sistema}'
        print('Servico não encontrado')
        return None

    def update_sistema(self, novo_nome):
        if self.get_sistema():
            for linha in range(2, ws.max_row + 1):
                if ws.cell(row=linha, column=1).value == self.codigo:
                    ws.cell(row=linha, column=2).value = novo_nome
                    wb.save(filename=filename)
            print('Sistema Atualizado!')
        else:
            print('Sistema não encontrado, impossivel atualizar')

    def delete_sistema(self):
        if self.get_sistema():
            for linha in range(2, ws.max_row + 1):
                if ws.cell(row=linha, column=1).value == self.codigo:
                    # ws.cell(row=linha, column=3).value = 1
                    ws.delete_rows(linha)
                    wb.save(filename=filename)
            print('Sistema Deletado!')
        else:
            print('Sistema não encontrado, impossivel deletar')

    def __str__(self):
        return f'Código => {self.codigo} Nome => {self.sistema}'



